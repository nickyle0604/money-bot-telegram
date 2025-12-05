import time
import requests
from openpyxl import load_workbook

from market_utils import get_price
from telegram_utils import send_telegram_message

# ===== CẤU HÌNH =====
BOT_TOKEN = "8119671883:AAFWhDPhryZnmLKv6RZmvbiMUItEo3H9hKc"
CHAT_ID = 1333105427   # có thể bỏ nếu muốn dùng chat_id động

API_URL = f"https://api.telegram.org/bot{BOT_TOKEN}"
SYMBOLS = ["BTCUSDT", "ETHUSDT", "LINKUSDT", "BNBUSDT"]
REPORT_FILE = "p2p_report.xlsx"


def get_updates(offset=None):
    params = {"timeout": 30}
    if offset is not None:
        params["offset"] = offset
    r = requests.get(f"{API_URL}/getUpdates", params=params, timeout=35)
    data = r.json()
    return data.get("result", [])


# ====== HÀM XỬ LÝ DỮ LIỆU P2P ======

def read_pnl_from_excel(filename=REPORT_FILE):
    """Đọc file Excel p2p_report.xlsx và trả về tổng PnL + 5 lệnh gần nhất."""
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        return None, []

    ws = wb.active
    total_pnl = 0.0
    rows = []

    # Giả sử cấu trúc: Amount | Buy | Sell | PnL (giống file bạn export)
    for row in ws.iter_rows(min_row=2, values_only=True):
        amount, buy, sell, pnl = row
        if pnl is None:
            continue
        total_pnl += float(pnl)
        rows.append((amount, buy, sell, pnl))

    last_rows = rows[-5:] if len(rows) > 5 else rows
    return total_pnl, last_rows


def format_history(last_rows):
    if not last_rows:
        return "Chưa có lịch sử lệnh trong file Excel."
    lines = ["5 lệnh gần nhất:"]
    for i, (amount, buy, sell, pnl) in enumerate(last_rows, 1):
        lines.append(
            f"{i}. {amount} USDT | Buy {buy} | Sell {sell} | PnL: {pnl}"
        )
    return "\n".join(lines)


# ====== HÀM XỬ LÝ COMMAND ======

def handle_command(chat_id, text):
    text = text.strip()

    if text.startswith("/start"):
        msg = (
            "Xin chào 👋\n"
            "Bot hỗ trợ các lệnh:\n"
            "/pnl - Xem tổng lãi/lỗ trong Excel\n"
            "/monitor - Xem giá hiện tại các coin\n"
            "/history - Xem 5 lệnh gần nhất\n"
        )
        send_telegram_message(BOT_TOKEN, chat_id, msg)

    elif text.startswith("/pnl"):
        total_pnl, _ = read_pnl_from_excel()
        if total_pnl is None:
            send_telegram_message(
                BOT_TOKEN, chat_id,
                "Không tìm thấy file p2p_report.xlsx trong thư mục dự án."
            )
        else:
            send_telegram_message(
                BOT_TOKEN, chat_id,
                f"TỔNG LÃI/LỖ (từ Excel): {total_pnl:.2f} USDT"
            )

    elif text.startswith("/history"):
        _, last_rows = read_pnl_from_excel()
        msg = format_history(last_rows)
        send_telegram_message(BOT_TOKEN, chat_id, msg)

    elif text.startswith("/monitor"):
        lines = ["Giá hiện tại:"]

        for sym in SYMBOLS:
            price = get_price(sym)
            if price is None:
                lines.append(f"{sym}: lỗi lấy giá")
            else:
                lines.append(f"{sym}: {price:.4f}")

        send_telegram_message(BOT_TOKEN, chat_id, "\n".join(lines))

    else:
        send_telegram_message(
            BOT_TOKEN, chat_id,
            "Không hiểu lệnh.\nDùng /start để xem danh sách lệnh."
        )


def main():
    offset = None
    print("Bot command đang chạy... Nhắn /start trong Telegram để test.")
    while True:
        updates = get_updates(offset)
        for u in updates:
            offset = u["update_id"] + 1

            message = u.get("message") or u.get("edited_message")
            if not message:
                continue

            chat = message.get("chat", {})
            chat_id = chat.get("id")
            text = message.get("text") or ""

            if not text or not chat_id:
                continue

            print("Nhận lệnh:", chat_id, text)
            handle_command(chat_id, text)

        time.sleep(1)


if __name__ == "__main__":
    main()
