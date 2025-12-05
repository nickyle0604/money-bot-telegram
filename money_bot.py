import time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path

# ========== CẤU HÌNH ==========
BOT_TOKEN = "8523051917:AAE76dwvNPXgxK3gHJcS4ZSFIyeKW3Pkm8o"   # <-- DÁN TOKEN BOT VÀO ĐÂY
CHAT_ID = 1333105427                    # <-- CHAT_ID của bạn

API_URL = f"https://api.telegram.org/bot{BOT_TOKEN}"
EXCEL_FILE = "money_log.xlsx"


# ========== FORMAT TIỀN ==========
def format_vnd(amount: float) -> str:
    """Format tiền kiểu VND: 1.234.567"""
    try:
        value = round(float(amount))
    except Exception:
        value = 0
    return f"{value:,}".replace(",", ".")


# ========== TELEGRAM ==========
def send_telegram_message(text: str):
    """Gửi tin nhắn về Telegram (1 chat cố định)."""
    try:
        requests.post(
            f"{API_URL}/sendMessage",
            json={"chat_id": CHAT_ID, "text": text},
            timeout=10
        )
    except Exception as e:
        print("Lỗi gửi Telegram:", e)


def get_updates(offset=None):
    """Lấy update mới từ Telegram."""
    params = {"timeout": 30}
    if offset is not None:
        params["offset"] = offset

    try:
        r = requests.get(f"{API_URL}/getUpdates", params=params, timeout=35)
        data = r.json()
        return data.get("result", [])
    except Exception as e:
        print("Lỗi get_updates:", e)
        return []


# ========== EXCEL ==========
def init_excel():
    """Tạo file Excel nếu chưa có."""
    if Path(EXCEL_FILE).exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Money"
    ws.append(["Date", "Type", "Amount", "Note", "Balance"])
    wb.save(EXCEL_FILE)
    print("Đã tạo file:", EXCEL_FILE)


def read_last_balance() -> float:
    """Lấy số dư cuối cùng. Nếu chưa có thì 0."""
    if not Path(EXCEL_FILE).exists():
        return 0.0

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    if ws.max_row < 2:
        return 0.0

    value = ws.cell(ws.max_row, 5).value
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def add_transaction(tran_type: str, date_str: str, amount: float, note: str) -> float:
    """
    Ghi 1 giao dịch mới vào Excel.
    tran_type: "IN" (tiền vào) hoặc "OUT" (tiền ra)
    date_str: dạng dd/mm/yyyy
    amount: số tiền dương
    note: ghi chú
    Trả về: số dư mới
    """
    init_excel()

    # xử lý ngày
    try:
        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
    except Exception:
        date_obj = datetime.now()

    date_save = date_obj.strftime("%Y-%m-%d")

    # tính số dư mới
    old_balance = read_last_balance()
    new_balance = old_balance + amount if tran_type == "IN" else old_balance - amount

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([date_save, tran_type, amount, note, new_balance])
    wb.save(EXCEL_FILE)

    return new_balance


def get_last_transactions(n=5):
    """Lấy n giao dịch gần nhất."""
    if not Path(EXCEL_FILE).exists():
        return []

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return rows[-n:] if len(rows) >= n else rows


# ========== PARSE INPUT ==========
def parse_input(text: str):
    """
    Định dạng lệnh:
      + SỐ_TiỀN NỘI_DUNG NGÀY
      - SỐ_TiỀN NỘI_DUNG NGÀY

    Ví dụ:
      + 1500000 Luong 5
      - 200000 Cafe 6
      + 2500000 Thu_nhap 12/11/2024

    NGÀY:
      - Nếu chỉ nhập 1–2 số (vd: 5) -> dùng ngày đó trong THÁNG/NĂM hiện tại
      - Nếu nhập dd/mm/yyyy -> dùng đúng ngày đó
    """
    parts = text.split()

    if len(parts) < 4:
        return None, None, None, "Sai định dạng. Ví dụ: + 150000 Luong 5"

    # parts[0] = + hoặc -
    amount_str = parts[1]
    day_input = parts[-1]            # ngày luôn là từ cuối cùng
    note = " ".join(parts[2:-1])     # nội dung = từ thứ 3 -> trước ngày

    today = datetime.now()

    # ===== Xử lý ngày =====
    if day_input.isdigit() and 1 <= len(day_input) <= 2:
        # chỉ nhập số ngày -> gắn tháng + năm hiện tại
        try:
            date_obj = datetime(today.year, today.month, int(day_input))
            date_str = date_obj.strftime("%d/%m/%Y")
        except ValueError:
            return None, None, None, "Ngày không hợp lệ."
    else:
        # yêu cầu dạng dd/mm/yyyy
        try:
            datetime.strptime(day_input, "%d/%m/%Y")
            date_str = day_input
        except ValueError:
            return None, None, None, "Ngày phải là dd/mm/yyyy hoặc chỉ nhập số ngày."

    # ===== Xử lý số tiền =====
    clean = amount_str.replace(".", "").replace(",", "")
    try:
        amount = float(clean)
    except ValueError:
        return None, None, None, "Số tiền không hợp lệ."

    return date_str, amount, note, None


# ========== HELP ==========
HELP_TEXT = (
    "📒 BOT GHI CHÉP TIỀN NHANH\n\n"
    "Lệnh:\n"
    "+ SỐ_TiỀN NỘI_DUNG NGÀY    → TIỀN VÀO\n"
    "- SỐ_TiỀN NỘI_DUNG NGÀY    → TIỀN RA\n"
    "=                           → XEM SỐ DƯ\n"
    ">                           → VỊ TRÍ FILE EXCEL\n"
    "/last                       → 5 GIAO DỊCH GẦN NHẤT\n"
    "/help hoặc /start           → XEM HƯỚNG DẪN\n\n"
    "NGÀY:\n"
    "  • Chỉ nhập số ngày (vd: 5) -> dùng ngày đó trong THÁNG/NĂM hiện tại\n"
    "  • Nếu khác tháng/năm      -> nhập dd/mm/yyyy (vd: 12/11/2024)\n\n"
    "Ví dụ:\n"
    "  + 1500000 Luong 5\n"
    "  - 200000 Cafe 6\n"
    "  + 2500000 Thu_nhap 12/11/2024\n"
)


# ========== XỬ LÝ LỆNH ==========
def handle_command(chat_id: int, text: str):
    if chat_id != CHAT_ID:
        return

    text = text.strip()

    # HELP
    if text in ("/start", "/help"):
        send_telegram_message(HELP_TEXT)
        return

    # XEM SỐ DƯ
    if text == "=":
        bal = read_last_balance()
        send_telegram_message(f"💰 Số dư: {format_vnd(bal)} VND")
        return

    # VỊ TRÍ FILE EXCEL
    if text == ">":
        path = Path(EXCEL_FILE).resolve()
        send_telegram_message(f"📁 File Excel: {path}")
        return

    # 5 GIAO DỊCH GẦN NHẤT
    if text == "/last":
        rows = get_last_transactions(5)
        if not rows:
            send_telegram_message("Chưa có giao dịch nào.")
            return

        lines = ["📄 5 giao dịch gần nhất:"]
        for d, t, a, n, b in rows:
            lines.append(
                f"{d} | {t} | {format_vnd(a)} | {n} | SD: {format_vnd(b)}"
            )
        send_telegram_message("\n".join(lines))
        return

    # TIỀN VÀO: "+"
    if text.startswith("+"):
        date_str, amount, note, err = parse_input(text)
        if err:
            send_telegram_message(err)
            return

        new_bal = add_transaction("IN", date_str, amount, note)
        msg = (
            "➕ TIỀN VÀO\n"
            f"Ngày: {date_str}\n"
            f"Tiền: {format_vnd(amount)} VND\n"
            f"Ghi chú: {note}\n"
            f"Số dư mới: {format_vnd(new_bal)} VND"
        )
        send_telegram_message(msg)
        return

    # TIỀN RA: "-"
    if text.startswith("-"):
        date_str, amount, note, err = parse_input(text)
        if err:
            send_telegram_message(err)
            return

        new_bal = add_transaction("OUT", date_str, amount, note)
        msg = (
            "➖ TIỀN RA\n"
            f"Ngày: {date_str}\n"
            f"Tiền: {format_vnd(amount)} VND\n"
            f"Ghi chú: {note}\n"
            f"Số dư mới: {format_vnd(new_bal)} VND"
        )
        send_telegram_message(msg)
        return

    # KHÔNG HIỂU
    send_telegram_message("Không hiểu lệnh. Gõ /help để xem cú pháp.")


# ========== MAIN LOOP ==========
def main():
    init_excel()
    offset = None
    print("Money bot đang chạy... Gõ /start trong Telegram.")

    while True:
        updates = get_updates(offset)

        for u in updates:
            offset = u["update_id"] + 1

            msg = u.get("message") or u.get("edited_message")
            if not msg:
                continue

            chat = msg.get("chat", {})
            chat_id = chat.get("id")
            text = msg.get("text") or ""

            if not text:
                continue

            print("Lệnh:", chat_id, text)
            handle_command(chat_id, text)

        time.sleep(1)


if __name__ == "__main__":
    main()


