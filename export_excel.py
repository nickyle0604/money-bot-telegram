import openpyxl
from openpyxl.styles import Font, PatternFill

def export_to_excel(data, filename="p2p_report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P2P Orders"

    headers = ["Amount", "Buy", "Sell", "PnL"]
    ws.append(headers)

    for row in data:
        ws.append(row)

    # Format màu
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value > 0:
                cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")  # xanh
            else:
                cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")  # đỏ

    wb.save(filename)
    print(f"Đã xuất file Excel: {filename}")
