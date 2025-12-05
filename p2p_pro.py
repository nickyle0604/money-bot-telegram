from tabulate import tabulate
import openpyxl
from openpyxl.styles import PatternFill
import requests

orders = []  # mỗi phần tử: [amount, buy, sell, pnl]


# ================== PHẦN PNL CƠ BẢN ==================

def calculate_pnl(amount, buy_price, sell_price):
    return (sell_price - buy_price) * amount


def add_order():
    print("\n=== NHẬP LỆNH P2P MỚI ===")
    amount = float(input("Số lượng (USDT): "))
    buy = float(input("Giá mua: "))
    sell = float(input("Giá bán: "))

    pnl = calculate_pnl(amount, buy, sell)
    orders.append([amount, buy, sell, pnl])
    print(f"Đã thêm lệnh, PnL = {pnl:.2f}\n")


def color_pnl(value):
    """Đổi PnL thành chuỗi có màu (xanh/lãi, đỏ/lỗ) cho terminal."""
    if value > 0:
        return f"\033[92m{value:.2f}\033[0m"  # xanh
    elif value < 0:
        return f"\033[91m{value:.2f}\033[0m"  # đỏ
    else:
        return f"{value:.2f}"


def show_orders():
    if not orders:
        print("\nChưa có lệnh nào.\n")
        return

    print("\n=== DANH SÁCH LỆNH ===")
    headers = ["Amount", "Buy", "Sell", "PnL"]

    table = []
    for amt, buy, sell, pnl in orders:
        table.append([amt, buy, sell, color_pnl(pnl)])

    print(tabulate(table, headers=headers, floatfmt=".2f"))

    # Thống kê
    total_pnl = sum(o[3] for o in orders)
    wins = sum(1 for o in orders if o[3] > 0)
    losses = sum(1 for o in orders if o[3] < 0)
    win_rate = (wins / len(orders)) * 100 if orders else 0

    print(f"\nTỔNG LÃI/LỖ: {total_pnl:.2f}")
    print(f"WIN: {wins} lệnh")
    print(f"LOSE: {losses} lệnh")
    print(f"TỶ LỆ WIN: {win_rate:.2f}%")

    # PnL lũy kế
    print("\nLÃI LŨY KẾ:")
    cumulative = 0
    for o in orders:
        cumulative += o[3]
        print(f"{cumulative:.2f}")
    print("")


# ================== PHẦN XUẤT EXCEL ==================

def export_to_excel(filename="p2p_report.xlsx"):
    if not orders:
        print("\nChưa có lệnh nào để xuất Excel.\n")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P2P Orders"

    headers = ["Amount", "Buy", "Sell", "PnL"]
    ws.append(headers)

    for amt, buy, sell, pnl in orders:
        ws.append([amt, buy, sell, pnl])

    # tô màu cột PnL
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value is None:
                continue
            if cell.value > 0:
                cell.fill = PatternFill(start_color="C6EFCE",
                                        end_color="C6EFCE",
                                        fill_type="solid")  # xanh
            elif cell.value < 0:
                cell.fill = PatternFill(start_color="FFC7CE",
                                        end_color="FFC7CE",
                                        fill_type="solid")  # đỏ

    wb.save(filename)
    print(f"\n✅ Đã xuất file Excel: {filename}\n")


# ================== PHẦN THỊ TRƯỜNG: GIÁ & ORDERBOOK ==================

BINANCE_BASE = "https://api.binance.com"


def get_link_price():
    url = f"{BINANCE_BASE}/api/v3/ticker/price?symbol=LINKUSDT"
    r = requests.get(url, timeout=5)
    data = r.json()
    price = float(data["price"])
    print(f"\nGiá LINK/USDT hiện tại: {price:.4f}\n")
    return price


def price_alert(old_price):
    print(f"\nGiá tham chiếu: {old_price:.4f}")
    new_price = get_link_price()
    change = ((new_price - old_price) / old_price) * 100
    print(f"Thay đổi: {change:.2f}%")

    if abs(change) > 1:
        print("⚠️  ALERT: LINK biến động > 1% !!!\n")
    else:
        print("✅ Biến động < 1%, vẫn trong vùng an toàn.\n")


def get_orderbook():
    url = f"{BINANCE_BASE}/api/v3/depth?symbol=LINKUSDT&limit=50"
    r = requests.get(url, timeout=5)
    data = r.json()

    buy_volume = sum(float(b[1]) for b in data["bids"])
    sell_volume = sum(float(a[1]) for a in data["asks"])

    print("\n=== PHÂN TÍCH ORDERBOOK LINK/USDT ===")
    print(f"Tổng volume BUY:  {buy_volume:.2f}")
    print(f"Tổng volume SELL: {sell_volume:.2f}")

    if buy_volume > sell_volume:
        print("→ ÁP LỰC MUA MẠNH (thiên hướng TĂNG)\n")
    elif sell_volume > buy_volume:
        print("→ ÁP LỰC BÁN MẠNH (thiên hướng GIẢM)\n")
    else:
        print("→ CÂN BẰNG, khó đoán xu hướng.\n")


# ================== MENU CHÍNH ==================

def main_menu():
    last_price = None

    while True:
        print("========== P2P PRO TOOL ==========")
        print("1. Nhập lệnh P2P")
        print("2. Xem bảng lệnh + thống kê")
        print("3. Xuất Excel")
        print("4. Xem giá LINK realtime")
        print("5. Cảnh báo biến động > 1% (so với giá tham chiếu)")
        print("6. Phân tích áp lực mua/bán (orderbook)")
        print("0. Thoát")
        choice = input("Chọn chức năng: ").strip()

        if choice == "1":
            add_order()
        elif choice == "2":
            show_orders()
        elif choice == "3":
            export_to_excel()
        elif choice == "4":
            last_price = get_link_price()
        elif choice == "5":
            if last_price is None:
                try:
                    ref_price = float(input("Nhập giá tham chiếu (VD: giá mua, hoặc giá cũ): "))
                except ValueError:
                    print("Giá không hợp lệ.\n")
                    continue
            else:
                use_last = input("Dùng giá LINK vừa xem gần nhất làm tham chiếu? (y/n): ").lower()
                if use_last == "y":
                    ref_price = last_price
                else:
                    try:
                        ref_price = float(input("Nhập giá tham chiếu: "))
                    except ValueError:
                        print("Giá không hợp lệ.\n")
                        continue
            price_alert(ref_price)
        elif choice == "6":
            get_orderbook()
        elif choice == "0":
            print("Thoát tool. Hẹn gặp lại!")
            break
        else:
            print("Lựa chọn không hợp lệ.\n")


if __name__ == "__main__":
    main_menu()
